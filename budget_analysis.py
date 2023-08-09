import pandas as pd
import os
from datetime import datetime
from tabulate import tabulate
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

# Folder names
SCALE_INSIGHTS_FOLDER = 'ScaleInsights'
BULK_FILES_FOLDER = 'bulk files'

# File names
SCALE_INSIGHTS_FILENAME_INPUT = 'AdGroupStats.csv'
BULK_FILES_FILENAME_INPUT = 'bulk_file.xlsx'
FILENAME_OUTPUT_SUFFIX = '_output'

# Full paths (input and output)
SCALE_INSIGHTS_INPUT_PATH = os.path.join(SCALE_INSIGHTS_FOLDER, SCALE_INSIGHTS_FILENAME_INPUT)
BULK_FILES_INPUT_PATH = os.path.join(BULK_FILES_FOLDER, BULK_FILES_FILENAME_INPUT)
SCALE_INSIGHTS_OUTPUT_PATH = os.path.join(SCALE_INSIGHTS_FOLDER, FILENAME_OUTPUT_SUFFIX + SCALE_INSIGHTS_FILENAME_INPUT)
BULK_FILES_OUTPUT_PATH = os.path.join(BULK_FILES_FOLDER, FILENAME_OUTPUT_SUFFIX + BULK_FILES_FILENAME_INPUT)

def create_folders():
    os.makedirs(SCALE_INSIGHTS_FOLDER, exist_ok=True)
    os.makedirs(BULK_FILES_FOLDER, exist_ok=True)

configs = {
    "SP": {
        "MIN_SPEND": 1.5,
        "MIN_SPEND_MANUAL_RANKING": 3.5,
    },
    "SB": {
        "MIN_SPEND": 1,
        "MAX_SPEND": 3.5,
    },
    "SD": {
        "MIN_SPEND": 1,
        "MAX_SPEND": 3.5,
    },
    "GENERAL": {
        "click_threshold_tester": 5,
        "MIN_SPEND_TESTER_CAMPAIGNS": 3.0,
        "NUM_DAYS": 14,
        "WEIGHTING_FACTOR": 1.5
    }
}

def find_column(sheet, target_header):
    for cell in sheet[1]:
        if cell.value == target_header:
            return cell.column

def update_bulk_file(df):
    workbook = load_workbook(BULK_FILES_INPUT_PATH)
    budget_mapping = df.set_index('Campaign')['daily_spend'].to_dict()
    sheets_and_budget_headers = {
        'Sponsored Products Campaigns': 'Daily Budget',
        'Sponsored Brands Campaigns': 'Budget',
        'Sponsored Display Campaigns': 'Budget'
    }

    for sheet_name, budget_header in sheets_and_budget_headers.items():
        sheet = workbook[sheet_name]
        entity_col = find_column(sheet, 'Entity')
        operation_col = find_column(sheet, 'Operation')
        campaign_name_col = find_column(sheet, 'Campaign Name')
        budget_col = find_column(sheet, budget_header)

        for row in sheet.iter_rows(min_row=2):
            if row[entity_col-1].value == 'Campaign':
                row[operation_col-1].value = 'Update'
                campaign_name = row[campaign_name_col-1].value
                if campaign_name in budget_mapping:
                    row[budget_col-1].value = budget_mapping[campaign_name]
    
    workbook.save(BULK_FILES_OUTPUT_PATH) # Save in the 'bulk files' directory
    
def load_and_preprocess_data():
    
    df = pd.read_csv(SCALE_INSIGHTS_INPUT_PATH)

    df.fillna(0, inplace=True)
    df = df.drop(columns=['Units', 'ROAS', 'Default Bid'])
    return df

def apply_constraints(df):
    df = df[df['State'] != 'Archived'].copy()
    df['%ad spend'] = df['%ad spend'].round(4)
    df['%ad sales'] = df['%ad sales'].round(4)
    df['distributed_spend'] = df['distributed_spend'].round(2)
    df['daily_spend'] = df['daily_spend'].round(2)

    for ctype in ['SP', 'SB', 'SD']:
        cfg = configs[ctype]
        rows = df['Type'].str.startswith(ctype)
        df.loc[rows, 'daily_spend'] = df.loc[rows, 'daily_spend'].apply(lambda x: max(x, cfg['MIN_SPEND']))
        if ctype in ['SB', 'SD']:
            df.loc[rows, 'daily_spend'] = df.loc[rows, 'daily_spend'].apply(lambda x: min(x, cfg['MAX_SPEND']))

    df.loc[(df['Type'] == 'SP Manual') & (df['AdGroup'] == 'Ranking'), 'daily_spend'] = df.loc[
        (df['Type'] == 'SP Manual') & (df['AdGroup'] == 'Ranking'), 'daily_spend'].apply(lambda x: max(x, configs["SP"]["MIN_SPEND_MANUAL_RANKING"]))

    df.loc[(df['Clicks'] < configs["GENERAL"]["click_threshold_tester"]) & (df['Type'].str.contains('SP')), 'daily_spend'] = df[
        'daily_spend'].apply(lambda x: max(x, configs["GENERAL"]["MIN_SPEND_TESTER_CAMPAIGNS"]))

    return df

def calculate_metrics(df):
    total_spend = df['Spent'].sum()
    total_sales = df['Sales'].sum()

    df['%ad spend'] = df['Spent'] / total_spend
    df['%ad sales'] = df['Sales'] / total_sales

    # Apply the weighting factor to campaigns with the AdGroup "Ranking"
    df['weighted_ad_sales'] = df.apply(lambda x: x['%ad sales'] * configs['GENERAL']['WEIGHTING_FACTOR'] if x['AdGroup'] == 'Ranking' else x['%ad sales'], axis=1)
    
    # Distribute spend using the weighted percentage
    df['distributed_spend'] = total_spend * df['weighted_ad_sales']
    df['daily_spend'] = df['distributed_spend'] / configs["GENERAL"]["NUM_DAYS"]

    return df

def calculate_and_print_results(df):
    daily_tester_budget = df.loc[(df['Clicks'] < configs["GENERAL"]["click_threshold_tester"]) & (df['Type'].str.contains('SP')), 'daily_spend'].sum()
    daily_other_budget = df.loc[~((df['Clicks'] < configs["GENERAL"]["click_threshold_tester"]) & (df['Type'].str.contains('SP'))), 'daily_spend'].sum()
    total_daily_budget = df['daily_spend'].sum()

    total_spend = df['Spent'].sum()
    total_initial_spend_daily = round(total_spend/configs['GENERAL']['NUM_DAYS'], 2)
    percent_daily_tester_budget = (daily_tester_budget/total_initial_spend_daily)*100
    percent_daily_other_budget = (daily_other_budget/total_initial_spend_daily)*100
    percent_total_daily_budget = (total_daily_budget/total_initial_spend_daily)*100

    results = {
        "Total initial spend (daily)": f"£{total_initial_spend_daily}",
        "Daily budget for tester campaigns": f"£{daily_tester_budget} ({percent_daily_tester_budget:.2f}%)",
        "Daily budget for other campaigns": f"£{daily_other_budget} ({percent_daily_other_budget:.2f}%)",
        "Total daily budget": f"£{total_daily_budget} ({percent_total_daily_budget:.2f}%)"
    }

    print(tabulate(results.items(), tablefmt='plain'))


def main():

    df = load_and_preprocess_data()
    df = calculate_metrics(df)
    df = apply_constraints(df)
    calculate_and_print_results(df)
    update_bulk_file(df)
    df.to_csv(SCALE_INSIGHTS_OUTPUT_PATH, index=False) # Save in the 'ScaleInsights' directory

if __name__ == "__main__":
    main()

